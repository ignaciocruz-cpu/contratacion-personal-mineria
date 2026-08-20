# -*- coding: utf-8 -*-
"""
Plataforma de Reclutamiento - Flesan Minería
Backend FastAPI con SQLite y almacenamiento local de documentos.
"""

from contextlib import asynccontextmanager
from fastapi import FastAPI, UploadFile, File, HTTPException, Request
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
import sqlite3
import shutil
import io
import re
import unicodedata
import openpyxl
from pathlib import Path
from datetime import datetime
import uvicorn, threading, webbrowser, time

@asynccontextmanager
async def lifespan(_app: FastAPI):
    init_db()
    print("[INFO] Base de datos lista.")
    yield

app = FastAPI(title="Flesan Minería - Reclutamiento", lifespan=lifespan)
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

HERE    = Path(__file__).parent
UPLOADS = HERE / "uploads"
DB_PATH = HERE / "flesan_recluta.db"
UPLOADS.mkdir(exist_ok=True)

# ── Base de datos ───────────────────────────────────────────────────────────────

def _norm_sql(s):
    """Función registrada en SQLite: minúsculas sin tildes."""
    if s is None:
        return ""
    import unicodedata as _ud, re as _re
    s = _ud.normalize("NFD", str(s).lower().strip())
    s = "".join(c for c in s if _ud.category(c) != "Mn")
    return _re.sub(r"\s+", " ", s)

def get_db():
    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    conn.create_function("norm", 1, _norm_sql)
    return conn

def init_db():
    conn = get_db()
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS candidatos (
            id                    INTEGER PRIMARY KEY AUTOINCREMENT,
            apellido_paterno      TEXT,
            apellido_materno      TEXT,
            primer_nombre         TEXT,
            segundo_nombre        TEXT,
            rut                   TEXT UNIQUE,
            fecha_nacimiento      TEXT,
            estado_civil          TEXT,
            nacionalidad          TEXT DEFAULT 'Chilena',
            telefono              TEXT,
            correo                TEXT,
            direccion             TEXT,
            comuna                TEXT,
            region                TEXT,
            cargo                 TEXT,
            profesion             TEXT,
            experiencia_general   INTEGER DEFAULT 0,
            experiencia_especifica INTEGER DEFAULT 0,
            talla_overol          TEXT,
            talla_zapatos         TEXT,
            licencia_conducir     TEXT,
            categoria             TEXT,
            especialidad          TEXT,
            notas                 TEXT,
            created_at            TEXT DEFAULT (datetime('now','localtime')),
            updated_at            TEXT DEFAULT (datetime('now','localtime'))
        );

        CREATE TABLE IF NOT EXISTS documentos (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            candidato_id  INTEGER NOT NULL,
            tipo          TEXT NOT NULL,
            filename      TEXT NOT NULL,
            filepath      TEXT NOT NULL,
            uploaded_at   TEXT DEFAULT (datetime('now','localtime')),
            FOREIGN KEY (candidato_id) REFERENCES candidatos(id) ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS ofertas (
            id               INTEGER PRIMARY KEY AUTOINCREMENT,
            titulo           TEXT NOT NULL,
            cargo            TEXT,
            profesion        TEXT,
            categoria        TEXT,
            especialidad     TEXT,
            experiencia_min  INTEGER DEFAULT 0,
            region           TEXT,
            descripcion      TEXT,
            activa           INTEGER DEFAULT 1,
            created_at       TEXT DEFAULT (datetime('now','localtime'))
        );

        CREATE TABLE IF NOT EXISTS pipeline (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            candidato_id INTEGER NOT NULL,
            oferta_id    INTEGER NOT NULL,
            etapa        TEXT NOT NULL DEFAULT 'Postulado',
            score        REAL DEFAULT 0,
            created_at   TEXT DEFAULT (datetime('now','localtime')),
            updated_at   TEXT DEFAULT (datetime('now','localtime')),
            FOREIGN KEY (candidato_id) REFERENCES candidatos(id) ON DELETE CASCADE,
            FOREIGN KEY (oferta_id)    REFERENCES ofertas(id)    ON DELETE CASCADE,
            UNIQUE(candidato_id, oferta_id)
        );

        CREATE TABLE IF NOT EXISTS actividad (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            candidato_id INTEGER NOT NULL,
            tipo         TEXT NOT NULL DEFAULT 'nota',
            descripcion  TEXT,
            created_at   TEXT DEFAULT (datetime('now','localtime')),
            FOREIGN KEY (candidato_id) REFERENCES candidatos(id) ON DELETE CASCADE
        );
    """)
    # Migraciones para columnas nuevas en tablas existentes
    for col, ddl in [
        ("estado",      "TEXT DEFAULT 'Disponible'"),
        ("creado_por",  "TEXT DEFAULT ''"),
        ("origen",      "TEXT DEFAULT 'manual'"),
    ]:
        try:
            conn.execute(f"ALTER TABLE candidatos ADD COLUMN {col} {ddl}")
        except Exception:
            pass
    for col, ddl in [("creado_por", "TEXT DEFAULT ''")]:
        try:
            conn.execute(f"ALTER TABLE ofertas ADD COLUMN {col} {ddl}")
        except Exception:
            pass
    conn.commit()
    conn.close()

# ── Static ──────────────────────────────────────────────────────────────────────

@app.get("/")
def root():
    return FileResponse(HERE / "index.html")

@app.get("/logo-mineria.png")
def logo_blanco():
    return FileResponse(HERE / "logo-mineria.png")

@app.get("/logo-mineria-color.png")
def logo_color():
    return FileResponse(HERE / "logo-mineria-color.png")

# ── Candidatos ──────────────────────────────────────────────────────────────────

CAMPOS_CANDIDATO = [
    "apellido_paterno", "apellido_materno", "primer_nombre", "segundo_nombre",
    "rut", "fecha_nacimiento", "estado_civil", "nacionalidad",
    "telefono", "correo", "direccion", "comuna", "region",
    "cargo", "profesion", "experiencia_general", "experiencia_especifica",
    "talla_overol", "talla_zapatos", "licencia_conducir",
    "categoria", "especialidad", "notas", "estado", "creado_por",
]

@app.get("/api/candidatos")
def listar_candidatos(
    q: str = "", cargo: str = "", region: str = "",
    exp_min: int = 0, categoria: str = "", especialidad: str = "",
    limit: int = 100, offset: int = 0,
):
    conn = get_db()
    c = conn.cursor()
    sql = """SELECT c.*, COALESCE(d.cnt, 0) as n_docs, COALESCE(dt.tipos, '') as doc_tipos
             FROM candidatos c
             LEFT JOIN (SELECT candidato_id, COUNT(*) cnt FROM documentos GROUP BY candidato_id) d
             ON c.id = d.candidato_id
             LEFT JOIN (SELECT candidato_id, GROUP_CONCAT(tipo) tipos FROM documentos GROUP BY candidato_id) dt
             ON c.id = dt.candidato_id
             WHERE 1=1"""
    params = []

    if q:
        p = f"%{_norm_sql(q)}%"
        sql += """ AND (
            norm(c.primer_nombre)    LIKE ? OR
            norm(c.apellido_paterno) LIKE ? OR
            norm(c.apellido_materno) LIKE ? OR
            norm(c.segundo_nombre)   LIKE ? OR
            norm(c.rut)              LIKE ? OR
            norm(c.correo)           LIKE ? OR
            norm(c.cargo)            LIKE ? OR
            norm(c.profesion)        LIKE ? OR
            norm(c.especialidad)     LIKE ? OR
            norm(c.comuna)           LIKE ?
        )"""
        params.extend([p] * 10)
    if cargo:
        sql += " AND norm(c.cargo) LIKE ?"
        params.append(f"%{_norm_sql(cargo)}%")
    if region:
        sql += " AND c.region = ?"
        params.append(region)
    if exp_min:
        sql += " AND CAST(c.experiencia_general AS INTEGER) >= ?"
        params.append(exp_min)
    if categoria:
        sql += " AND c.categoria = ?"
        params.append(categoria)
    if especialidad:
        sql += " AND norm(c.especialidad) LIKE ?"
        params.append(f"%{_norm_sql(especialidad)}%")

    sql += " ORDER BY c.apellido_paterno, c.primer_nombre"

    # Total sin paginación
    count_sql = f"SELECT COUNT(*) FROM ({sql})"
    c.execute(count_sql, params)
    total = c.fetchone()[0]

    sql += " LIMIT ? OFFSET ?"
    c.execute(sql, params + [limit, offset])
    rows = [dict(r) for r in c.fetchall()]

    conn.close()
    return {"total": total, "candidatos": rows}

@app.get("/api/candidatos/exportar")
def exportar_candidatos():
    from openpyxl.styles import Font, PatternFill, Alignment
    conn = get_db()
    c = conn.cursor()
    c.execute("""SELECT c.*, COALESCE(d.cnt,0) as n_docs, COALESCE(dt.tipos,'') as doc_tipos
                 FROM candidatos c
                 LEFT JOIN (SELECT candidato_id, COUNT(*) cnt FROM documentos GROUP BY candidato_id) d
                   ON c.id = d.candidato_id
                 LEFT JOIN (SELECT candidato_id, GROUP_CONCAT(tipo) tipos FROM documentos GROUP BY candidato_id) dt
                   ON c.id = dt.candidato_id
                 ORDER BY c.apellido_paterno, c.primer_nombre""")
    rows = [dict(r) for r in c.fetchall()]
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Candidatos"

    headers = [
        "ID","Apellido Paterno","Apellido Materno","Primer Nombre","Segundo Nombre","RUT",
        "Fecha Nacimiento","Estado Civil","Nacionalidad","Teléfono","Correo",
        "Dirección","Comuna","Región","Cargo","Profesión",
        "Exp. General (años)","Exp. Específica (años)",
        "Talla Overol","Talla Zapatos","Licencia Conducir","Categoría","Especialidad",
        "Estado","Notas","N° Documentos","Tipos Documentos","Creado Por","Fecha Registro","Última Actualización",
    ]
    fields = [
        "id","apellido_paterno","apellido_materno","primer_nombre","segundo_nombre","rut",
        "fecha_nacimiento","estado_civil","nacionalidad","telefono","correo",
        "direccion","comuna","region","cargo","profesion",
        "experiencia_general","experiencia_especifica",
        "talla_overol","talla_zapatos","licencia_conducir","categoria","especialidad",
        "estado","notas","n_docs","doc_tipos","creado_por","created_at","updated_at",
    ]

    fill = PatternFill("solid", fgColor="E30613")
    font_h = Font(bold=True, color="FFFFFF", size=10)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = fill
        cell.font = font_h
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    for r_idx, row in enumerate(rows, 2):
        for c_idx, field in enumerate(fields, 1):
            ws.cell(row=r_idx, column=c_idx, value=row.get(field))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fecha = datetime.now().strftime("%Y%m%d_%H%M")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=candidatos_{fecha}.xlsx"},
    )

@app.get("/api/candidatos/{cid}")
def get_candidato(cid: int):
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM candidatos WHERE id = ?", (cid,))
    row = c.fetchone()
    if not row:
        conn.close()
        raise HTTPException(404, "Candidato no encontrado")
    d = dict(row)
    c.execute("SELECT * FROM documentos WHERE candidato_id = ? ORDER BY tipo", (cid,))
    d["documentos"] = [dict(r) for r in c.fetchall()]
    conn.close()
    return d

@app.post("/api/candidatos")
async def crear_candidato(request: Request):
    data = await request.json()
    vals = {k: data.get(k) for k in CAMPOS_CANDIDATO}
    vals["origen"] = "postulacion" if data.get("origen") == "postulacion" else "manual"
    cols = ", ".join(vals.keys())
    ph   = ", ".join(["?"] * len(vals))
    conn = get_db()
    try:
        cur = conn.execute(f"INSERT INTO candidatos ({cols}) VALUES ({ph})", list(vals.values()))
        new_id = cur.lastrowid
        conn.execute("INSERT INTO actividad (candidato_id,tipo,descripcion) VALUES (?,?,?)",
                     (new_id, "registro", "Perfil registrado en el sistema"))
        conn.commit()
        conn.close()
        return {"ok": True, "id": new_id}
    except sqlite3.IntegrityError:
        conn.close()
        raise HTTPException(400, "El RUT ya existe en la base de datos")

@app.put("/api/candidatos/{cid}")
async def actualizar_candidato(cid: int, request: Request):
    data = await request.json()
    vals = {k: data[k] for k in CAMPOS_CANDIDATO if k in data}
    vals["updated_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    set_clause = ", ".join(f"{k} = ?" for k in vals)
    conn = get_db()
    try:
        if "estado" in vals:
            c2 = conn.cursor()
            c2.execute("SELECT estado FROM candidatos WHERE id=?", (cid,))
            row2 = c2.fetchone()
            if row2 and row2[0] != vals["estado"]:
                conn.execute("INSERT INTO actividad (candidato_id,tipo,descripcion) VALUES (?,?,?)",
                             (cid, "estado", f"Estado cambiado a '{vals['estado']}'"))
        conn.execute(f"UPDATE candidatos SET {set_clause} WHERE id = ?", list(vals.values()) + [cid])
        conn.commit()
        conn.close()
        return {"ok": True}
    except sqlite3.IntegrityError:
        conn.close()
        raise HTTPException(400, "El RUT ya existe en otro candidato")

@app.delete("/api/candidatos/{cid}")
def eliminar_candidato(cid: int):
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT filepath FROM documentos WHERE candidato_id = ?", (cid,))
    for row in c.fetchall():
        Path(row["filepath"]).unlink(missing_ok=True)
    cand_dir = UPLOADS / str(cid)
    shutil.rmtree(cand_dir, ignore_errors=True)
    conn.execute("DELETE FROM candidatos WHERE id = ?", (cid,))
    conn.commit()
    conn.close()
    return {"ok": True}

# ── Documentos ───────────────────────────────────────────────────────────────────

TIPOS_DOC = {
    "cv":                  "CV",
    "cert_titulo":         "Certificado de Título",
    "cert_afp":            "Certificado AFP",
    "cert_salud":          "Certificado Salud",
    "comprobante_domicilio": "Comprobante Domicilio",
    "carnet":              "Fotocopia Carnet",
    "transferencia":       "Datos Transferencia",
    "hoja_vida_conductor": "Hoja de Vida Conductor",
}

@app.post("/api/candidatos/{cid}/documentos/{tipo}")
async def subir_documento(cid: int, tipo: str, file: UploadFile = File(...)):
    if tipo not in TIPOS_DOC:
        raise HTTPException(400, f"Tipo inválido: {tipo}")
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT id FROM candidatos WHERE id = ?", (cid,))
    if not c.fetchone():
        conn.close()
        raise HTTPException(404, "Candidato no encontrado")

    cand_dir = UPLOADS / str(cid)
    cand_dir.mkdir(exist_ok=True)
    ext = Path(file.filename).suffix
    filepath = cand_dir / f"{tipo}{ext}"

    with open(filepath, "wb") as f:
        shutil.copyfileobj(file.file, f)

    c.execute("SELECT id FROM documentos WHERE candidato_id = ? AND tipo = ?", (cid, tipo))
    existing = c.fetchone()
    if existing:
        conn.execute(
            "UPDATE documentos SET filename=?, filepath=?, uploaded_at=datetime('now','localtime') WHERE id=?",
            (file.filename, str(filepath), existing["id"])
        )
    else:
        conn.execute(
            "INSERT INTO documentos (candidato_id, tipo, filename, filepath) VALUES (?,?,?,?)",
            (cid, tipo, file.filename, str(filepath))
        )
    conn.commit()
    conn.close()
    return {"ok": True, "filename": file.filename}

@app.get("/api/documentos/{doc_id}/descargar")
def descargar_documento(doc_id: int):
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM documentos WHERE id = ?", (doc_id,))
    doc = c.fetchone()
    conn.close()
    if not doc:
        raise HTTPException(404, "Documento no encontrado")
    fp = Path(doc["filepath"])
    if not fp.exists():
        raise HTTPException(404, "Archivo no encontrado en disco")
    return FileResponse(str(fp), filename=doc["filename"])

@app.delete("/api/documentos/{doc_id}")
def eliminar_documento(doc_id: int):
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT filepath FROM documentos WHERE id = ?", (doc_id,))
    doc = c.fetchone()
    if doc:
        Path(doc["filepath"]).unlink(missing_ok=True)
        conn.execute("DELETE FROM documentos WHERE id = ?", (doc_id,))
        conn.commit()
    conn.close()
    return {"ok": True}

# ── Ofertas ──────────────────────────────────────────────────────────────────────

@app.get("/api/ofertas")
def listar_ofertas(q: str = ""):
    conn = get_db()
    c = conn.cursor()
    if q:
        p = f"%{_norm_sql(q)}%"
        c.execute("""SELECT * FROM ofertas
                     WHERE norm(titulo) LIKE ? OR norm(cargo) LIKE ?
                        OR norm(profesion) LIKE ? OR norm(especialidad) LIKE ?
                        OR norm(descripcion) LIKE ?
                     ORDER BY created_at DESC""",
                  [p, p, p, p, p])
    else:
        c.execute("SELECT * FROM ofertas ORDER BY created_at DESC")
    rows = [dict(r) for r in c.fetchall()]
    conn.close()
    return {"ofertas": rows}

@app.post("/api/ofertas")
async def crear_oferta(request: Request):
    d = await request.json()
    conn = get_db()
    cur = conn.execute(
        "INSERT INTO ofertas (titulo,cargo,profesion,categoria,especialidad,experiencia_min,region,descripcion,creado_por) VALUES (?,?,?,?,?,?,?,?,?)",
        (d.get("titulo",""), d.get("cargo",""), d.get("profesion",""),
         d.get("categoria",""), d.get("especialidad",""),
         d.get("experiencia_min",0), d.get("region",""), d.get("descripcion",""),
         d.get("creado_por",""))
    )
    conn.commit()
    conn.close()
    return {"ok": True, "id": cur.lastrowid}

CAMPOS_OFERTA = ["titulo","cargo","profesion","categoria","especialidad",
                  "experiencia_min","region","descripcion","activa"]

@app.put("/api/ofertas/{oid}")
async def actualizar_oferta(oid: int, request: Request):
    d = await request.json()
    vals = {k: d[k] for k in CAMPOS_OFERTA if k in d}
    if not vals:
        return {"ok": True}
    set_clause = ", ".join(f"{k} = ?" for k in vals)
    conn = get_db()
    conn.execute(f"UPDATE ofertas SET {set_clause} WHERE id = ?", list(vals.values()) + [oid])
    conn.commit()
    conn.close()
    return {"ok": True}

@app.delete("/api/ofertas/{oid}")
def eliminar_oferta(oid: int):
    conn = get_db()
    conn.execute("DELETE FROM ofertas WHERE id = ?", (oid,))
    conn.commit()
    conn.close()
    return {"ok": True}

# ── Motor de Matching v2 ──────────────────────────────────────────────────────────

def _norm(s: str) -> str:
    """Normaliza: minúsculas, sin tildes, espacios simples."""
    s = unicodedata.normalize("NFD", (s or "").lower().strip())
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", s)

# Grupos de sinónimos para cargos en minería / industria
_GRUPOS_CARGO = [
    {"operador de planta","operador de equipos","operador","operador de maquinaria","operador de proceso"},
    {"tecnico electricista","electricista industrial","tecnico en electricidad","electricista"},
    {"mecanico industrial","tecnico mecanico","mecanico de mantenimiento","mecanico","oficial mecanico"},
    {"supervisor de terreno","supervisor de area","jefe de turno","supervisor","capataz"},
    {"soldador","soldador certificado","tecnico soldador","soldador 6g"},
    {"tecnico en instrumentacion","instrumentista","tecnico instrumentista","tecnico en automatizacion"},
    {"topografo","tecnico en topografia","asistente topografo"},
    {"perforista","operador de perforacion","tronador","operador de jumbos"},
    {"bodeguero","encargado de bodega","panolero","auxiliar de bodega"},
    {"ingeniero de minas","ingeniero en minas","ingeniero civil de minas"},
    {"operador de grua","gruero","operador grua horquilla","operador grua horquilla"},
    {"oficial de mantencion","oficial de mantenimiento","tecnico de mantencion","tecnico mantenimiento"},
    {"operador de cargador frontal","operador cargador frontal","operador de retroexcavadora"},
    {"jefe de proyecto","administrador de contrato","coordinador de proyecto"},
    {"prevencionista","prevencionista de riesgos","tecnico en prevencion de riesgos"},
]

def _cargo_score(a: str, b: str) -> float:
    """Similitud entre dos cargos: 0.0 – 1.0"""
    na, nb = _norm(a), _norm(b)
    if not na or not nb: return 0.0
    if na == nb: return 1.0
    if na in nb or nb in na: return 0.82
    for grupo in _GRUPOS_CARGO:
        if na in grupo and nb in grupo: return 0.88
    wa = {w for w in na.split() if len(w) > 3}
    wb = {w for w in nb.split() if len(w) > 3}
    if wa and wb:
        ov = len(wa & wb) / max(len(wa), len(wb))
        if ov >= 0.5: return 0.50
        if ov > 0:    return 0.25
    return 0.0

def _text_score(a: str, b: str) -> float:
    """Similitud genérica entre dos textos: 0.0 – 1.0"""
    na, nb = _norm(a), _norm(b)
    if not na or not nb: return 0.0
    if na == nb: return 1.0
    if na in nb or nb in na: return 0.72
    wa = {w for w in na.split() if len(w) > 3}
    wb = {w for w in nb.split() if len(w) > 3}
    if wa and wb:
        ov = len(wa & wb) / max(len(wa), len(wb))
        if ov >= 0.5: return 0.48
        if ov > 0:    return 0.22
    return 0.0

def _exp_score(candidato: int, requerido: int) -> float:
    """Score continuo de experiencia (no binario)."""
    if requerido <= 0: return 1.0
    if candidato <= 0: return 0.0
    ratio = candidato / requerido
    if ratio >= 1.5: return 0.93   # sobre-calificado → leve reducción
    if ratio >= 1.0: return 1.0
    if ratio >= 0.80: return 0.75
    if ratio >= 0.60: return 0.45
    if ratio >= 0.40: return 0.20
    return 0.05

# Pesos fijos de cada dimensión (suman 100)
_PESOS = {"cargo": 35, "profesion": 20, "exp": 18, "especialidad": 12, "region": 10, "categoria": 5}

def _score_v2(cand: dict, oferta: dict) -> dict:
    dims = {}
    detalle = []

    # 1. Cargo (35 pts) — siempre evaluado
    sc = _cargo_score(oferta.get("cargo"), cand.get("cargo"))
    dims["cargo"] = {"pts": round(sc * 35, 1), "max": 35, "label": "Cargo / Rol"}
    if sc >= 0.95:   detalle.append("Cargo exacto")
    elif sc >= 0.80: detalle.append("Cargo muy similar")
    elif sc >= 0.45: detalle.append("Cargo relacionado")

    # 2. Profesión (20 pts)
    if oferta.get("profesion"):
        sp = _text_score(oferta.get("profesion"), cand.get("profesion"))
        dims["profesion"] = {"pts": round(sp * 20, 1), "max": 20, "label": "Profesión"}
        if sp >= 0.90: detalle.append("Profesión exacta")
        elif sp >= 0.45: detalle.append("Profesión relacionada")
    else:
        dims["profesion"] = {"pts": 20, "max": 20, "label": "Profesión", "na": True}

    # 3. Experiencia General (18 pts)
    exp_req = int(oferta.get("experiencia_min") or 0)
    exp_can = int(cand.get("experiencia_general") or 0)
    if exp_req > 0:
        se = _exp_score(exp_can, exp_req)
        dims["exp"] = {"pts": round(se * 18, 1), "max": 18, "label": "Experiencia"}
        if se >= 1.0:    detalle.append(f"{exp_can}a exp ✓ (req. {exp_req}a)")
        elif se >= 0.70: detalle.append(f"{exp_can}a exp (req. {exp_req}a)")
        else:            detalle.append(f"Exp. insuficiente ({exp_can}a / {exp_req}a)")
    else:
        dims["exp"] = {"pts": 18, "max": 18, "label": "Experiencia", "na": True}

    # 4. Especialidad (12 pts) — match exacto sobre valores controlados
    if oferta.get("especialidad"):
        ne = _norm(oferta.get("especialidad", ""))
        nc_e = _norm(cand.get("especialidad", ""))
        ss = 1.0 if ne and nc_e and ne == nc_e else (0.5 if ne and nc_e and (ne in nc_e or nc_e in ne) else 0.0)
        dims["especialidad"] = {"pts": round(ss * 12, 1), "max": 12, "label": "Especialidad"}
        if ss >= 1.0: detalle.append("Especialidad exacta")
        elif ss > 0:  detalle.append("Especialidad compatible")
    else:
        dims["especialidad"] = {"pts": 12, "max": 12, "label": "Especialidad", "na": True}

    # 5. Región (10 pts)
    if oferta.get("region"):
        sr = 1.0 if _norm(oferta.get("region", "")) == _norm(cand.get("region", "")) else 0.0
        dims["region"] = {"pts": round(sr * 10, 1), "max": 10, "label": "Región"}
        if sr == 1.0: detalle.append("Misma región")
        else:         detalle.append("Región diferente")
    else:
        dims["region"] = {"pts": 10, "max": 10, "label": "Región", "na": True}

    # 6. Categoría (5 pts) — match exacto sobre valores controlados
    if oferta.get("categoria"):
        nc = _norm(oferta.get("categoria", ""))
        cc2 = _norm(cand.get("categoria", ""))
        scat = 1.0 if nc and cc2 and nc == cc2 else 0.0
        dims["categoria"] = {"pts": round(scat * 5, 1), "max": 5, "label": "Categoría"}
        if scat: detalle.append("Categoría exacta")
        else:    detalle.append("Categoría distinta")
    else:
        dims["categoria"] = {"pts": 5, "max": 5, "label": "Categoría", "na": True}

    score_base = round(sum(d["pts"] for d in dims.values()), 1)

    # Bonus perfil completo (hasta 5 pts)
    bonus = 0.0
    if cand.get("correo"):     bonus += 1.0
    if cand.get("telefono"):   bonus += 1.0
    if cand.get("profesion"):  bonus += 0.5
    if cand.get("especialidad"): bonus += 0.5
    n_docs = int(cand.get("n_docs") or 0)
    bonus += min(2.0, n_docs * 0.5)
    bonus = round(min(5.0, bonus), 1)

    score = min(100.0, round(score_base + bonus, 1))

    return {"score": score, "score_base": score_base, "bonus_perfil": bonus,
            "dimensiones": dims, "detalle": detalle}

@app.get("/api/ofertas/{oid}/match")
def matching(oid: int):
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM ofertas WHERE id = ?", (oid,))
    oferta = c.fetchone()
    if not oferta:
        conn.close()
        raise HTTPException(404, "Oferta no encontrada")
    oferta = dict(oferta)

    c.execute("""SELECT c.*, COALESCE(d.cnt,0) as n_docs
                 FROM candidatos c
                 LEFT JOIN (SELECT candidato_id, COUNT(*) cnt FROM documentos GROUP BY candidato_id) d
                 ON c.id = d.candidato_id""")
    candidatos = [dict(r) for r in c.fetchall()]
    conn.close()

    resultados, excluidos = [], []
    exp_req_global = int(oferta.get("experiencia_min") or 0)

    for cand in candidatos:
        # — Exclusión dura 1: cargo completamente distinto —
        if oferta.get("cargo") and _cargo_score(oferta.get("cargo"), cand.get("cargo")) == 0:
            excluidos.append({**cand, "razon_exclusion": "Cargo no compatible"})
            continue
        # — Exclusión dura 2: experiencia < 35% del mínimo requerido —
        if exp_req_global > 0:
            exp_can = int(cand.get("experiencia_general") or 0)
            if exp_can < exp_req_global * 0.35:
                excluidos.append({**cand, "razon_exclusion": f"Exp. insuficiente ({exp_can}a / mín. {exp_req_global}a)"})
                continue

        m = _score_v2(cand, oferta)
        if m["score"] >= 15:
            resultados.append({**cand, **m})

    resultados.sort(key=lambda x: x["score"], reverse=True)
    total = len(resultados)
    for i, r in enumerate(resultados):
        r["rank"] = i + 1
        r["percentil"] = round((1 - i / total) * 100) if total > 1 else 100

    return {"oferta": oferta, "matches": resultados, "total": total,
            "excluidos": len(excluidos), "total_evaluados": len(candidatos)}

# ── Carga masiva ─────────────────────────────────────────────────────────────────

EXCEL_COL_MAP = {
    "apellido paterno":            "apellido_paterno",
    "apellido materno":            "apellido_materno",
    "primer nombre":               "primer_nombre",
    "segundo nombre":              "segundo_nombre",
    "rut":                         "rut",
    "fecha de nacimiento":         "fecha_nacimiento",
    "estado civil":                "estado_civil",
    "nacionalidad":                "nacionalidad",
    "teléfono":                    "telefono",
    "telefono":                    "telefono",
    "correo electrónico":          "correo",
    "correo electronico":          "correo",
    "correo":                      "correo",
    "dirección":                   "direccion",
    "direccion":                   "direccion",
    "comuna":                      "comuna",
    "región":                      "region",
    "region":                      "region",
    "cargo":                       "cargo",
    "profesión":                   "profesion",
    "profesion":                   "profesion",
    "experiencia general (años)":  "experiencia_general",
    "experiencia general (anos)":  "experiencia_general",
    "experiencia general":         "experiencia_general",
    "experiencia específica (años)": "experiencia_especifica",
    "experiencia especifica (anos)": "experiencia_especifica",
    "experiencia específica":      "experiencia_especifica",
    "experiencia especifica":      "experiencia_especifica",
    "talla overol":                "talla_overol",
    "talla zapatos":               "talla_zapatos",
    "licencia conducir":           "licencia_conducir",
    "categoría":                   "categoria",
    "categoria":                   "categoria",
    "especialidad":                "especialidad",
    "notas":                       "notas",
    "estado":                      "estado",
    # Variantes abreviadas del exportador
    "fecha nacimiento":            "fecha_nacimiento",
    "exp. general (años)":         "experiencia_general",
    "exp. general (anos)":         "experiencia_general",
    "exp. específica (años)":      "experiencia_especifica",
    "exp. especifica (anos)":      "experiencia_especifica",
    "exp. específica (anos)":      "experiencia_especifica",
}

@app.post("/api/candidatos/bulk/preview")
async def preview_masivo(file: UploadFile = File(...)):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Solo se aceptan archivos .xlsx o .xls")
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(400, "No se pudo leer el archivo Excel")
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(400, "El archivo está vacío")
    headers = [str(h) if h is not None else "" for h in rows[0]]
    data_rows = [[str(v) if v is not None else None for v in row] for row in rows[1:] if any(v is not None for v in row)]
    return {"headers": headers, "rows": data_rows, "total": len(data_rows)}


@app.post("/api/candidatos/bulk")
async def carga_masiva(file: UploadFile = File(...)):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Solo se aceptan archivos .xlsx o .xls")

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(400, "No se pudo leer el archivo Excel")

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(400, "El archivo no tiene datos")

    # Mapear encabezados
    raw_headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    col_map = {}  # índice → campo DB
    for i, h in enumerate(raw_headers):
        db_field = EXCEL_COL_MAP.get(h.lower())
        if db_field:
            col_map[i] = db_field

    if "rut" not in col_map.values():
        raise HTTPException(400, "El archivo no tiene columna RUT")

    conn = get_db()
    importados, omitidos, errores = 0, 0, []

    for row_num, row in enumerate(rows[1:], start=2):
        vals = {}
        for i, db_field in col_map.items():
            v = row[i] if i < len(row) else None
            vals[db_field] = str(v).strip() if v is not None and str(v).strip() != "" else None

        if not vals.get("rut"):
            omitidos += 1
            continue

        # Solo insertar campos válidos de candidatos
        insert_vals = {k: v for k, v in vals.items() if k in CAMPOS_CANDIDATO}
        insert_vals["origen"] = "carga_masiva"
        cols = ", ".join(insert_vals.keys())
        ph   = ", ".join(["?"] * len(insert_vals))
        try:
            conn.execute(f"INSERT INTO candidatos ({cols}) VALUES ({ph})", list(insert_vals.values()))
            importados += 1
        except sqlite3.IntegrityError:
            omitidos += 1
        except Exception as e:
            errores.append(f"Fila {row_num}: {str(e)}")

    conn.commit()
    conn.close()
    return {"importados": importados, "omitidos": omitidos, "errores": errores}


@app.get("/api/plantilla")
def descargar_plantilla():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Candidatos"
    headers = [
        "Apellido Paterno","Apellido Materno","Primer Nombre","Segundo Nombre","RUT",
        "Fecha de Nacimiento","Estado Civil","Nacionalidad","Teléfono","Correo Electrónico",
        "Dirección","Comuna","Región","Cargo","Profesión",
        "Experiencia General (años)","Experiencia Específica (años)",
        "Talla Overol","Talla Zapatos","Licencia Conducir","Categoría","Especialidad","Notas",
    ]
    from openpyxl.styles import Font, PatternFill, Alignment
    fill = PatternFill("solid", fgColor="E30613")
    font = Font(bold=True, color="FFFFFF", size=10)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
    ws.row_dimensions[1].height = 35
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=plantilla_candidatos.xlsx"},
    )


# ── Catálogos y Stats ─────────────────────────────────────────────────────────────

REGIONES = [
    "Región de Arica y Parinacota","Región de Tarapacá","Región de Antofagasta",
    "Región de Atacama","Región de Coquimbo","Región de Valparaíso",
    "Región Metropolitana","Región del Libertador B. O'Higgins","Región del Maule",
    "Región de Ñuble","Región del Biobío","Región de La Araucanía",
    "Región de Los Ríos","Región de Los Lagos","Región de Aysén",
    "Región de Magallanes",
]

@app.get("/api/catalogos")
def catalogos():
    conn = get_db()
    c = conn.cursor()
    def uniq(col):
        c.execute(f"SELECT DISTINCT {col} FROM candidatos WHERE {col} IS NOT NULL AND {col}!='' ORDER BY {col}")
        return [r[0] for r in c.fetchall()]
    result = {
        "cargos":        uniq("cargo"),
        "profesiones":   uniq("profesion"),
        "regiones":      REGIONES,
        "categorias":    ["Supervisor","MM","Capataz","M1","M2","TIG","MIG","Cañería","Calderería","HDP","Oxigenista","Alta","Baja"],
        "especialidades":["OOCC","EEII","Mecánico","Estructura","Piping","Soldador","Rigger"],
        "espec_categorias":{
            "OOCC":      ["Supervisor","Capataz","M1","M2"],
            "EEII":      ["Supervisor","MM","Capataz","M1","M2"],
            "Mecánico":  ["Supervisor","MM","Capataz","M1","M2"],
            "Estructura":["Supervisor","MM","Capataz","M1","M2"],
            "Piping":    ["Supervisor","MM","Capataz","M1","M2"],
            "Soldador":  ["TIG","MIG","Cañería","Calderería","HDP","Oxigenista"],
            "Rigger":    ["Alta","Baja"],
        },
        "estados_civiles":["Soltero/a","Casado/a","Divorciado/a","Viudo/a","Conviviente civil"],
        "tallas_overol": ["XS","S","M","L","XL","XXL","XXXL"],
        "licencias":     ["No","A","A1","A2","A3","A4","B","C","D","E","F"],
        "tipos_doc":     TIPOS_DOC,
    }
    conn.close()
    return result

@app.get("/api/stats")
def stats():
    from datetime import date
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT COUNT(*) FROM candidatos");                                        n_cand     = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM ofertas WHERE activa=1");                            n_ofertas  = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM documentos");                                        n_docs     = c.fetchone()[0]
    c.execute("SELECT COUNT(DISTINCT candidato_id) FROM documentos WHERE tipo='cv'");   n_cv       = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM candidatos WHERE correo IS NOT NULL AND correo!=''"); n_correo  = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM candidatos WHERE telefono IS NOT NULL AND telefono!=''"); n_tel = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM candidatos WHERE licencia_conducir IS NOT NULL AND licencia_conducir NOT IN ('','No')"); n_lic = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM candidatos WHERE estado='Disponible' OR estado IS NULL"); n_disp = c.fetchone()[0]
    mes = date.today().strftime("%Y-%m")
    c.execute("SELECT COUNT(*) FROM candidatos WHERE created_at LIKE ?", (f"{mes}%",)); n_mes = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM candidatos c WHERE NOT EXISTS (SELECT 1 FROM documentos d WHERE d.candidato_id=c.id)"); n_sin_docs = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM candidatos WHERE (correo IS NULL OR correo='') AND (telefono IS NULL OR telefono='')"); n_sin_contacto = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM candidatos WHERE origen='postulacion' AND created_at >= datetime('now','localtime','-1 day')"); n_postulaciones_nuevas = c.fetchone()[0]
    c.execute("SELECT cargo, COUNT(*) n FROM candidatos WHERE cargo!='' GROUP BY cargo ORDER BY n DESC LIMIT 6")
    top_cargos        = [dict(r) for r in c.fetchall()]
    c.execute("SELECT region, COUNT(*) n FROM candidatos WHERE region!='' GROUP BY region ORDER BY n DESC LIMIT 6")
    top_regiones      = [dict(r) for r in c.fetchall()]
    c.execute("SELECT especialidad, COUNT(*) n FROM candidatos WHERE especialidad!='' GROUP BY especialidad ORDER BY n DESC LIMIT 6")
    top_especialidades= [dict(r) for r in c.fetchall()]
    c.execute("SELECT categoria, COUNT(*) n FROM candidatos WHERE categoria!='' GROUP BY categoria ORDER BY n DESC")
    dist_categorias   = [dict(r) for r in c.fetchall()]
    c.execute("SELECT COALESCE(estado,'Disponible') estado, COUNT(*) n FROM candidatos GROUP BY estado ORDER BY n DESC")
    dist_estados      = [dict(r) for r in c.fetchall()]
    c.execute("SELECT nacionalidad, COUNT(*) n FROM candidatos WHERE nacionalidad!='' GROUP BY nacionalidad ORDER BY n DESC LIMIT 6")
    dist_nacionalidades=[dict(r) for r in c.fetchall()]
    conn.close()
    return {
        "n_candidatos": n_cand, "n_ofertas": n_ofertas, "n_documentos": n_docs,
        "n_cv": n_cv, "n_correo": n_correo, "n_telefono": n_tel,
        "n_con_licencia": n_lic, "n_disponibles": n_disp, "n_este_mes": n_mes,
        "n_sin_docs": n_sin_docs, "n_sin_contacto": n_sin_contacto,
        "n_postulaciones_nuevas": n_postulaciones_nuevas,
        "top_cargos": top_cargos, "top_regiones": top_regiones,
        "top_especialidades": top_especialidades, "dist_categorias": dist_categorias,
        "dist_estados": dist_estados, "dist_nacionalidades": dist_nacionalidades,
    }


# ── Pipeline ──────────────────────────────────────────────────────────────────────

ETAPAS_PIPELINE = ["Postulado","En Revisión","Entrevista","Evaluación","Seleccionado","Descartado"]

@app.get("/api/pipeline")
def get_pipeline(oferta_id: int = 0):
    conn = get_db()
    c = conn.cursor()
    base = """SELECT p.*, c.primer_nombre, c.apellido_paterno, c.apellido_materno,
                     c.cargo as cand_cargo, c.experiencia_general, c.region as cand_region,
                     o.titulo as oferta_titulo
              FROM pipeline p
              JOIN candidatos c ON p.candidato_id = c.id
              JOIN ofertas o    ON p.oferta_id    = o.id"""
    if oferta_id:
        c.execute(base + " WHERE p.oferta_id=? ORDER BY p.etapa, c.apellido_paterno", (oferta_id,))
    else:
        c.execute(base + " ORDER BY p.oferta_id, p.etapa, c.apellido_paterno")
    rows = [dict(r) for r in c.fetchall()]
    conn.close()
    return {"pipeline": rows, "etapas": ETAPAS_PIPELINE}

@app.post("/api/pipeline")
async def agregar_pipeline(request: Request):
    d = await request.json()
    conn = get_db()
    try:
        cur = conn.execute(
            "INSERT OR REPLACE INTO pipeline (candidato_id,oferta_id,etapa,score) VALUES (?,?,?,?)",
            (d["candidato_id"], d["oferta_id"], d.get("etapa","Postulado"), d.get("score",0))
        )
        conn.execute("INSERT INTO actividad (candidato_id,tipo,descripcion) VALUES (?,?,?)",
                     (d["candidato_id"], "pipeline", f"Agregado al pipeline: {d.get('oferta_titulo','')}"))
        conn.commit()
        conn.close()
        return {"ok": True, "id": cur.lastrowid}
    except Exception as e:
        conn.close()
        raise HTTPException(400, str(e))

@app.put("/api/pipeline/{pid}")
async def mover_pipeline(pid: int, request: Request):
    d = await request.json()
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT p.*,o.titulo FROM pipeline p JOIN ofertas o ON p.oferta_id=o.id WHERE p.id=?", (pid,))
    row = c.fetchone()
    if not row:
        conn.close()
        raise HTTPException(404)
    row = dict(row)
    nueva = d.get("etapa", row["etapa"])
    conn.execute("UPDATE pipeline SET etapa=?,updated_at=datetime('now','localtime') WHERE id=?", (nueva,pid))
    conn.execute("INSERT INTO actividad (candidato_id,tipo,descripcion) VALUES (?,?,?)",
                 (row["candidato_id"],"pipeline",f"Movido a '{nueva}' en '{row['titulo']}'"))
    conn.commit()
    conn.close()
    return {"ok": True}

@app.delete("/api/pipeline/{pid}")
def quitar_pipeline(pid: int):
    conn = get_db()
    conn.execute("DELETE FROM pipeline WHERE id=?", (pid,))
    conn.commit()
    conn.close()
    return {"ok": True}

# ── Actividad ─────────────────────────────────────────────────────────────────────

@app.get("/api/candidatos/{cid}/actividad")
def get_actividad(cid: int):
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM actividad WHERE candidato_id=? ORDER BY created_at DESC LIMIT 50", (cid,))
    rows = [dict(r) for r in c.fetchall()]
    conn.close()
    return {"actividad": rows}

@app.post("/api/candidatos/{cid}/actividad")
async def add_actividad(cid: int, request: Request):
    d = await request.json()
    conn = get_db()
    conn.execute("INSERT INTO actividad (candidato_id,tipo,descripcion) VALUES (?,?,?)",
                 (cid, d.get("tipo","nota"), d.get("descripcion","")))
    conn.commit()
    conn.close()
    return {"ok": True}

# ── Auto-match ───────────────────────────────────────────────────────────────────

@app.get("/api/candidatos/{cid}/auto-match")
def auto_match_candidato(cid: int):
    conn = get_db()
    c = conn.cursor()
    c.execute("""SELECT ca.*, COALESCE(d.cnt,0) as n_docs FROM candidatos ca
                 LEFT JOIN (SELECT candidato_id, COUNT(*) cnt FROM documentos GROUP BY candidato_id) d
                 ON ca.id=d.candidato_id WHERE ca.id=?""", (cid,))
    cand = c.fetchone()
    if not cand:
        conn.close()
        raise HTTPException(404)
    cand = dict(cand)
    c.execute("SELECT * FROM ofertas WHERE activa=1")
    ofertas = [dict(r) for r in c.fetchall()]
    conn.close()
    matches = []
    for o in ofertas:
        m = _score_v2(cand, o)
        if m["score"] >= 55:
            matches.append({"oferta_id": o["id"], "titulo": o["titulo"], "cargo": o["cargo"], "score": m["score"]})
    matches.sort(key=lambda x: x["score"], reverse=True)
    return {"matches": matches, "total": len(matches)}

# ── Formulario público ────────────────────────────────────────────────────────────

@app.get("/postulacion")
def postulacion():
    return FileResponse(HERE / "postulacion.html")

# ── Main ─────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    def _open():
        time.sleep(2)
        webbrowser.open("http://localhost:3025")
    threading.Thread(target=_open, daemon=True).start()
    uvicorn.run(app, host="0.0.0.0", port=3025, log_level="info")

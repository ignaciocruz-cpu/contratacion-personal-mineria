# -*- coding: utf-8 -*-
"""
Limpia y normaliza candidatos_20260820_1034.xlsx:
  1. RUT  : agrega puntos (18.574.590-2). Extranjeros quedan sin puntos.
  2. Teléfono: +569XXXXXXXX, elige el primero si hay dos números.
  3. Comunas: nombres oficiales estandarizados.
  4. Categoría y Especialidad: derivadas del Cargo.
Salida: candidatos_limpio.xlsx (no sobreescribe el original)
"""

import openpyxl, re, unicodedata
from pathlib import Path

SRC = Path(r'C:\Users\ignacio.cruz\OneDrive - Universidad de Chile\Escritorio\TRABAJO\06. Colina\11. Datos\12. Proyecto Jorge\candidatos_20260820_1034.xlsx')
DST = SRC.parent / 'candidatos_limpio.xlsx'


# ── Utilidades ────────────────────────────────────────────────────────────────

def _norm(s):
    """Mayúsculas sin tildes, sin símbolos ordinales."""
    s = unicodedata.normalize('NFD', str(s or '').upper())
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
    return s.replace('°', '').replace('ª', '').replace('º', '')


# ── 1. RUT ───────────────────────────────────────────────────────────────────

def format_rut(rut, nac):
    if not rut:
        return rut
    rut = str(rut).strip().upper().replace('.', '').replace(' ', '')
    if '-' in rut:
        num, dv = rut.rsplit('-', 1)
    else:
        num, dv = rut[:-1], rut[-1]
    if not num.isdigit():
        return rut  # Pasaporte u otro: no tocar
    nac_n = _norm(str(nac or ''))
    if nac_n not in ('', 'CHILENA', 'CHILENO'):
        return f"{num}-{dv}"  # Extranjero: sin puntos
    try:
        return f"{int(num):,}".replace(',', '.') + f"-{dv}"
    except Exception:
        return rut


# ── 2. TELÉFONO ──────────────────────────────────────────────────────────────

def format_phone(raw):
    if not raw or str(raw).strip() in ('', 'None', 'nan'):
        return ''
    s = str(raw).strip()

    # a) Separadores explícitos de múltiples números
    for sep in [' / ', '/', ';', ' – ', ' — ']:
        if sep in s:
            s = s.split(sep)[0].strip()
            break

    # b) Dos bloques de 7+ dígitos contiguos → tomar el primero
    grupos = re.findall(r'\d{7,}', s)
    if len(grupos) >= 2:
        s = grupos[0]

    # c) Solo dígitos
    digits = re.sub(r'[^\d]', '', s)

    # d) Quitar prefijos de país / área
    if digits.startswith('569') and len(digits) >= 11:
        digits = digits[3:]
    elif digits.startswith('56') and len(digits) >= 10:
        digits = digits[2:]
    elif digits.startswith('09') and len(digits) in (9, 10):
        digits = digits[1:]
    elif digits.startswith('0') and len(digits) == 9:
        digits = digits[1:]

    # e) Formatear
    if len(digits) == 9 and digits.startswith('9'):
        return f"+56{digits}"          # +56 9XXXXXXXX
    elif len(digits) == 8:
        return f"+569{digits}"         # +569 XXXXXXXX
    elif len(digits) > 9:
        cand = digits[-9:]
        if cand.startswith('9'):
            return f"+56{cand}"
        return f"+569{digits[-8:]}"
    elif len(digits) >= 7:
        return f"+569{digits.zfill(8)}"
    return raw  # Irreconocible: dejar original


# ── 3. COMUNAS ───────────────────────────────────────────────────────────────

COMUNAS = {
    'Chiguayante.':              'Chiguayante',
    'Chillan':                   'Chillán',
    'Con Con':                   'Concón',
    'Con-Con':                   'Concón',
    'Concon':                    'Concón',
    'Concepcion':                'Concepción',
    'Concepción - Talcahuano':   'Concepción',
    'Constitucion':              'Constitución',
    'Contitucion':               'Constitución',
    'Copiapo':                   'Copiapó',
    'Coronel.':                  'Coronel',
    'Curico':                    'Curicó',
    'Diego De Almagro':          'Diego de Almagro',
    'Hualpen':                   'Hualpén',
    'Hualpen.':                  'Hualpén',
    'Iquiique':                  'Iquique',
    'Laja.':                     'Laja',
    'Las Ventanas - Puchuncabí': 'Puchuncaví',
    'Llay Llay':                 'Llaillay',
    'Llay-Llay':                 'Llaillay',
    'Los Angeles':               'Los Ángeles',
    'Maipu':                     'Maipú',
    'Molina.':                   'Molina',
    'Quillon':                   'Quillón',
    'Quilpue':                   'Quilpué',
    'Romaico':                   'Renaico',
    'San Bernando':              'San Bernardo',
    'San Pedro De La Paz':       'San Pedro de la Paz',
    'San Pedro De La Paz.':      'San Pedro de la Paz',
    'Talca.':                    'Talca',
    'Tome':                      'Tomé',
    'V.Alemana':                 'Villa Alemana',
    'Valdiva':                   'Valdivia',
    'Valparaiso':                'Valparaíso',
    'Vilcun':                    'Vilcún',
    'Viña Del Mar':              'Viña del Mar',
}

def normalize_comuna(com):
    if not com:
        return com
    return COMUNAS.get(str(com).strip(), str(com).strip())


# ── 4. CARGO → CATEGORÍA / ESPECIALIDAD ─────────────────────────────────────

def cargo_a_cat_esp(cargo, cat_act, esp_act):
    if not cargo:
        return cat_act, esp_act
    c = _norm(cargo)

    # ── CATEGORÍA ──
    cat = None

    if 'MIG' in c and 'SOLDA' in c:
        cat = 'MIG'
    elif 'TIG' in c and 'SOLDA' in c:
        cat = 'TIG'
    elif 'SOLDA' in c:
        cat = 'Soldador'
    elif re.search(r'OXIGENIST|OXIGINISTA', c):
        cat = 'Oxigenista'
    elif 'RIGGER' in c:
        cat = 'Rigger'
    elif 'CAPATAZ' in c:
        cat = 'Capataz'
    elif 'SUPERVISOR' in c:
        cat = 'Supervisor'
    # MM
    elif re.search(
        r'(?:MAESTRO|MAETRO)\s+MAYOR'
        r'|M\.\s*MAYOR'
        r'|M\s*[-\.]\s*M\b'
        r'|-MM\b|\bMM\b'
        r'|\bM\s+MAYOR\b', c):
        cat = 'MM'
    # M1
    elif re.search(
        r'(?:MAESTRO|MAETRO)\s+PRIMERA'
        r'|(?:MAESTRO|MAETRO)\s+1'
        r'|M\.\s*PRIMERA'
        r'|1RA|1ERA'
        r'|-M1\b|\bM1\b'
        r'|\bM\s*[-.]?\s*1\b'
        r'|M\s{1,2}1\b', c):
        cat = 'M1'
    # M2
    elif re.search(
        r'(?:MAESTRO|MAETRO)\s+SEGUNDA'
        r'|(?:MAESTRO|MAETRO)\s+2'
        r'|M\.\s*SEGUNDA'
        r'|2DA'
        r'|-M2\b|\bM2\b'
        r'|\bM\s*[-.]?\s*2\b'
        r'|M\s{1,2}2\b'
        r'|MAYOR\s+SEGUNDA', c):
        cat = 'M2'

    if cat is None:
        cat = cat_act

    # ── ESPECIALIDAD ──
    esp = None

    if re.search(r'ELECTRICO|EEII|ELECTROMECANICO|INSTRUMENTAC', c):
        esp = 'EEII'
    elif re.search(r'PIPING|CANERIA|CANONERO', c):
        esp = 'Piping'
    elif re.search(r'MECANICO', c):
        esp = 'Mecánico'
    elif re.search(r'OBRAS\s+CIVILES|OOCC|OO\.CC|ALARIFE|ALBANIL', c):
        esp = 'OOCC'
    elif 'ESTRUCTURA' in c:
        esp = 'Estructura'
    elif re.search(r'ANDAMIO|RIGGER', c):
        esp = 'Rigger'
    elif re.search(r'SOLDA|OXIGENIST|OXIGINISTA', c):
        esp = 'Soldador'

    if esp is None:
        esp = esp_act

    return cat, esp


# ── PROCESO PRINCIPAL ─────────────────────────────────────────────────────────

wb = openpyxl.load_workbook(str(SRC))
ws = wb.active
headers = [cell.value for cell in ws[1]]

def col(name):
    n = _norm(name)
    for i, h in enumerate(headers):
        if _norm(str(h or '')) == n:
            return i
    return None

i_rut = col('RUT')
i_nac = col('Nacionalidad')
i_tel = col('Teléfono')
i_com = col('Comuna')
i_car = col('Cargo')
i_cat = col('Categoría')
i_esp = col('Especialidad')

print(f"Procesando {ws.max_row - 1} candidatos...")

cambios = {'rut': 0, 'tel': 0, 'com': 0, 'cat': 0, 'esp': 0}

for row in ws.iter_rows(min_row=2):
    v = [cell.value for cell in row]

    nac = v[i_nac] if i_nac is not None else 'Chilena'

    # RUT
    rut_new = format_rut(v[i_rut], nac)
    if rut_new and str(rut_new) != str(v[i_rut] or ''):
        row[i_rut].value = rut_new
        cambios['rut'] += 1

    # Teléfono
    if i_tel is not None:
        tel_new = format_phone(v[i_tel])
        if tel_new and str(tel_new) != str(v[i_tel] or ''):
            row[i_tel].value = tel_new
            cambios['tel'] += 1

    # Comuna
    com_new = normalize_comuna(v[i_com])
    if com_new and str(com_new) != str(v[i_com] or ''):
        row[i_com].value = com_new
        cambios['com'] += 1

    # Categoría y Especialidad
    cat_new, esp_new = cargo_a_cat_esp(v[i_car], v[i_cat], v[i_esp])
    if cat_new and str(cat_new) != str(v[i_cat] or ''):
        row[i_cat].value = cat_new
        cambios['cat'] += 1
    if esp_new and str(esp_new) != str(v[i_esp] or ''):
        row[i_esp].value = esp_new
        cambios['esp'] += 1

wb.save(str(DST))
print(f"\nArchivo guardado: {DST.name}")
print(f"\nCambios realizados:")
print(f"  RUT          : {cambios['rut']}")
print(f"  Teléfono     : {cambios['tel']}")
print(f"  Comuna       : {cambios['com']}")
print(f"  Categoría    : {cambios['cat']}")
print(f"  Especialidad : {cambios['esp']}")
print(f"  TOTAL        : {sum(cambios.values())}")

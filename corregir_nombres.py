# -*- coding: utf-8 -*-
"""
Corrige nombres y apellidos en candidatos_limpio.xlsx.

Se detectaron DOS patrones de ingreso incorrecto:

PATRÓN A (mayoritario, ~2100 filas):
  Columna "Apellido Paterno" (DB) → tiene el(los) NOMBRE(S) de pila
  Columna "Primer Nombre"  (DB) → tiene el APELLIDO PATERNO
  Columna "Segundo Nombre" (DB) → tiene el APELLIDO MATERNO
  → Se intercambian columnas.

PATRÓN B (~30-50 filas, ingresados correctamente):
  Columna "Apellido Paterno" (DB) → tiene el APELLIDO PATERNO correcto
  Columna "Primer Nombre"   (DB) → tiene el PRIMER NOMBRE correcto
  Columna "Segundo Nombre"  (DB) → tiene el APELLIDO MATERNO (pero en columna equivocada)
  → No se intercambia; solo se mueve S_NOM → Apellido Materno.

La distinción se hace chequeando si el valor actual de "Primer Nombre"
es un nombre de pila típico (→ Patrón B) o un apellido (→ Patrón A).

Salida: candidatos_final.xlsx + reporte_nombres.txt
"""

import openpyxl
import unicodedata
from pathlib import Path

SRC = Path(r'C:\Users\ignacio.cruz\OneDrive - Universidad de Chile\Escritorio\TRABAJO\06. Colina\11. Datos\12. Proyecto Jorge\candidatos_limpio.xlsx')
DST = SRC.parent / 'candidatos_final.xlsx'
RPT = SRC.parent / 'reporte_nombres.txt'


# ── Nombres de pila típicos chilenos ──────────────────────────────────────────
# Si el campo "Primer Nombre" (DB) contiene uno de estos → el registro NO está
# invertido (Patrón B): el apellido paterno ya está donde corresponde.
NOMBRES_PILA = {
    # Masculinos
    'abel','abner','abraham','adam','adan','adelio','adolfo','adrian','agustin',
    'alan','albert','alberto','alejandro','alex','alexis','alfredo','alvaro',
    'andres','angel','antonio','ariel','armando','arturo','augusto',
    'benjamin','bernardo','boris',
    'camilo','carlos','celso','cesar','christian','claudio','cristian','cristobal',
    'damian','daniel','darwin','david','diego',
    'edgardo','edmundo','eduardo','elias','emilio','enrique','eric','erick','ernesto',
    'esteban','eugenio','ezequiel',
    'fabian','felipe','felix','fernado','fernando','flavio','francisco','freddy','fredy',
    'gabriel','gerardo','gilberto','giovanni','gonzalo','guillermo','gustavo',
    'hector','hernan','hugo',
    'ignacio','ivan',
    'jaime','javier','jesus','jonathan','jorge','jose','julio',
    'kevin',
    'lazaro','leandro','leonel','leonardo','lorgio','lorenzo','lucas','luciano','luis',
    'manuel','marcelo','marco','mario','martin','matias','mauricio','maximiliano',
    'michael','miguel','moises',
    'nelson','nestor','nicolas','norman',
    'omar','oscar','osvaldo','oswaldo',
    'pablo','patricio','paulo','pedro',
    'rafael','ramiro','ramon','raul','rene','renato','ricardo','roberto',
    'rodrigo','rolando','roman','ruben',
    'samuel','saul','sebastian','segundo','sergio','silvio',
    'tomas',
    'ulises',
    'valentin','victor',
    'walter','william',
    'yonatan',
    # Femeninos
    'alejandra','andrea','ana','beatriz',
    'camila','carolina','catalina','cecilia','claudia',
    'daniela','diana',
    'elena','esperanza',
    'fabiola','fernanda','florencia','francisca',
    'gabriela','gisela','gladys','gloria','graciela',
    'irene','isabel',
    'javiera','jessica',
    'karen','karina',
    'laura','liliana','lorena','lucia','luisa',
    'magdalena','marcela','marcia','maria','mariela','marisa','marlene','martha',
    'melisa','miriam','mirna','monica',
    'noemi','norma',
    'olga',
    'pamela','patricia','patricia','paola','paula','paulina','pilar','priscilla',
    'raquel','rebeca','rocio','rosa','ruth',
    'sandra','silvia','sonia','sofia','susana',
    'tatiana','teresa','trinidad',
    'valentina','vanessa','veronica','viviana',
    'ximena',
    'yessenia','yolanda',
}

def _norm_key(s):
    """Minúsculas sin tildes para comparar en el set."""
    s = unicodedata.normalize('NFD', str(s or '').lower().strip())
    return ''.join(c for c in s if unicodedata.category(c) != 'Mn')

def is_nombre_pila(s):
    return _norm_key(s) in NOMBRES_PILA


# ── Title Case para nombres propios ──────────────────────────────────────────

PARTICULAS = {'de', 'del', 'la', 'las', 'los', 'el', 'y', 'e', 'al'}

def tc(s):
    """Title Case respetando partículas en posición no inicial."""
    if not s:
        return ''
    s = str(s).strip().lstrip(',').strip()
    palabras = s.split()
    result = []
    for i, p in enumerate(palabras):
        if i > 0 and p.lower() in PARTICULAS:
            result.append(p.lower())
        else:
            # capitalize respeta ñ, á, etc.
            result.append(p[0].upper() + p[1:].lower() if p else '')
    return ' '.join(result)


# ── Proceso ───────────────────────────────────────────────────────────────────

wb = openpyxl.load_workbook(str(SRC))
ws = wb.active
headers = [cell.value for cell in ws[1]]

def col(name):
    n = _norm_key(name)
    for i, h in enumerate(headers):
        if _norm_key(str(h or '')) == n:
            return i
    return None

i_ap = col('Apellido Paterno')
i_am = col('Apellido Materno')
i_pn = col('Primer Nombre')
i_sn = col('Segundo Nombre')
i_id = col('ID')

print(f"Procesando {ws.max_row - 1} filas...")

n_swap   = 0   # Patrón A: columnas invertidas → se intercambian
n_correc = 0   # Patrón B: ingresado correctamente → solo mover s_nom → ap_mat
n_sin_ap = 0   # Sin apellido: solo aplicar Title Case
alertas  = []

for row in ws.iter_rows(min_row=2):
    v   = [cell.value for cell in row]
    rid = v[i_id]
    ap  = str(v[i_ap] or '').strip().lstrip(',').strip()  # Apellido Paterno (DB)
    am  = str(v[i_am] or '').strip()                       # Apellido Materno (DB)
    pn  = str(v[i_pn] or '').strip()                       # Primer Nombre (DB)
    sn  = str(v[i_sn] or '').strip()                       # Segundo Nombre (DB)

    if not ap:
        # ── Sin apellido: solo Title Case en nombres ──────────────────────────
        if pn: row[i_pn].value = tc(pn)
        if sn: row[i_sn].value = tc(sn)
        n_sin_ap += 1
        continue

    if is_nombre_pila(pn):
        # ── Patrón B: datos en orden correcto ────────────────────────────────
        # Apellido Paterno ya está bien. Solo mover Segundo Nombre → Apellido Materno
        # si el segundo nombre parece un apellido (no un nombre de pila).
        row[i_ap].value = tc(ap)
        row[i_pn].value = tc(pn)
        if sn and not is_nombre_pila(sn):
            # El "segundo nombre" es en realidad el apellido materno
            row[i_am].value = tc(sn)
            row[i_sn].value = ''
        else:
            if sn: row[i_sn].value = tc(sn)
        n_correc += 1
    else:
        # ── Patrón A: columnas invertidas → intercambiar ──────────────────────
        # ap  = nombres de pila (1-3 palabras)
        # pn  = apellido paterno
        # sn  = apellido materno

        nuevo_ap_pat = tc(pn)   # apellido paterno real
        nuevo_ap_mat = tc(sn)   # apellido materno real

        # Separar los nombres del campo "ap"
        palabras = ap.split()
        if len(palabras) == 1:
            nuevo_pn = tc(palabras[0])
            nuevo_sn = ''
        elif len(palabras) == 2:
            nuevo_pn = tc(palabras[0])
            nuevo_sn = tc(palabras[1])
        else:
            # 3+ palabras: primera → primer nombre, resto → segundo nombre
            # Pero si la primera palabra NO es un nombre de pila, marcar alerta
            nuevo_pn = tc(palabras[0])
            nuevo_sn = tc(' '.join(palabras[1:]))
            if not is_nombre_pila(palabras[0]):
                alertas.append(
                    f"  ID {rid}: posible apellido en campo nombres — "
                    f"ap='{ap}' | pn='{pn}' | sn='{sn}'"
                )

        row[i_ap].value = nuevo_ap_pat
        row[i_am].value = nuevo_ap_mat
        row[i_pn].value = nuevo_pn
        row[i_sn].value = nuevo_sn
        n_swap += 1

wb.save(str(DST))

# ── Reporte ───────────────────────────────────────────────────────────────────
with open(str(RPT), 'w', encoding='utf-8') as f:
    f.write("REPORTE DE CORRECCIÓN DE NOMBRES Y APELLIDOS\n")
    f.write("=" * 60 + "\n\n")
    f.write(f"Filas con columnas intercambiadas (Patrón A): {n_swap}\n")
    f.write(f"Filas con datos correctos (Patrón B):         {n_correc}\n")
    f.write(f"Filas sin apellido (solo nombres):            {n_sin_ap}\n")
    f.write(f"TOTAL:                                        {n_swap+n_correc+n_sin_ap}\n\n")
    if alertas:
        f.write(f"CASOS PARA REVISIÓN MANUAL ({len(alertas)}):\n")
        f.write("-" * 60 + "\n")
        for a in alertas:
            f.write(a + "\n")

print(f"\nArchivo guardado: {DST.name}")
print(f"  Patrón A (swap)  : {n_swap}")
print(f"  Patrón B (correc): {n_correc}")
print(f"  Sin apellido     : {n_sin_ap}")
print(f"  Para revisión    : {len(alertas)}")
print(f"\nReporte: reporte_nombres.txt")

# -*- coding: utf-8 -*-
"""
Genera candidatos_ejemplo.xlsx con 47 personas ficticias para prueba de carga masiva.
Uso: python generar_excel.py
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import random
from datetime import date, timedelta

random.seed(42)

# ── Datos ficticios ──────────────────────────────────────────────────────────

PRIMER_NOMBRE_M = ["Rodrigo","Carlos","Felipe","Andrés","Diego","Sebastián","Cristián",
                   "Pablo","Luis","Eduardo","Ignacio","Alejandro","Nicolás","José",
                   "Francisco","Manuel","Gonzalo","Patricio","Ricardo","Daniel","Víctor"]
PRIMER_NOMBRE_F = ["María","Ana","Patricia","Claudia","Carolina","Valentina","Camila",
                   "Francisca","Lorena","Isabel","Sandra","Verónica","Cecilia","Andrea",
                   "Daniela","Javiera","Catalina","Fernanda","Sofía","Paola"]
SEGUNDO_NOMBRE  = ["Ignacio","Andrés","Antonio","Alejandro","Patricio","Alberto",
                   "María","Isabel","Andrea","Paz","Fernanda","Lorena",
                   "Enrique","Roberto","René","Marcelo","Cristóbal","","",""]

APELLIDOS = ["González","Muñoz","Rodríguez","López","Martínez","García","Pérez",
             "Sánchez","Ramírez","Torres","Flores","Rivera","Gómez","Díaz",
             "Morales","Vargas","Vásquez","Castro","Rojas","Fuentes","Hernández",
             "Espinoza","Contreras","Tapia","Salinas","Araya","Sepúlveda","Medina",
             "Ortiz","Leiva","Bravo","Campos","Mena","Cáceres","Pizarro","Vergara"]

CARGOS = ["Operador de Planta","Supervisor de Terreno","Técnico Electricista",
          "Mecánico Industrial","Jefe de Turno","Soldador","Operador de Grúa",
          "Ingeniero de Minas","Técnico en Instrumentación","Oficial de Mantención",
          "Bodeguero","Operador de Cargador Frontal","Perforista","Topógrafo",
          "Técnico en Electricidad"]

PROFESIONES = ["Técnico en Minería","Ingeniero Civil Industrial","Electricista Industrial",
               "Mecánico Industrial","Soldador Certificado","Geólogo","Ingeniero en Minas",
               "Topógrafo","Técnico Electrónico","Técnico en Mantención","Sin Profesión"]

REGIONES_COMUNAS = {
    "Región de Antofagasta":  ["Antofagasta","Calama","Tocopilla","Mejillones"],
    "Región de Atacama":      ["Copiapó","Caldera","Vallenar","Chañaral"],
    "Región de Coquimbo":     ["La Serena","Coquimbo","Ovalle","Illapel"],
    "Región Metropolitana":   ["Santiago","Pudahuel","Maipú","La Florida","Quilicura"],
    "Región de Valparaíso":   ["Valparaíso","Viña del Mar","Quillota","San Antonio"],
    "Región del Biobío":      ["Concepción","Talcahuano","Los Ángeles","Chillán"],
    "Región de La Araucanía": ["Temuco","Angol","Victoria","Villarrica"],
}

ESPECIALIDADES = ["EEII","Mecánica","Instrumentación","Soldadura","Civil","Mina","Proceso","Otro"]
CATEGORIAS     = ["MM","M1","M2","M3","Técnico","Operador","Profesional"]
ESTADOS_CIVILES= ["Soltero/a","Casado/a","Casado/a","Divorciado/a","Viudo/a"]
TALLAS_OVEROL  = ["S","M","M","L","L","XL","XXL"]
TALLAS_ZAPATOS = [str(t) for t in range(38, 46)]
LICENCIAS      = ["No","No","B","B","C","D","A2","A3"]
NACIONALIDADES = ["Chilena","Chilena","Chilena","Chilena","Chilena","Peruana","Boliviana","Colombiana"]

CALLES  = ["Av. Libertad","Calle Los Mineros","Pasaje El Cobre","Los Pinos","Av. Brasil",
           "Los Cerezos","Calle Atacama","Av. Condell","Calle Valdivia","Los Alamos"]


def rut_ficticio(n):
    """Genera RUT ficticio con dígito verificador correcto."""
    digits = [int(d) for d in str(n)]
    multipliers = [2, 3, 4, 5, 6, 7, 2, 3]
    s = sum(d * m for d, m in zip(reversed(digits), multipliers))
    r = 11 - (s % 11)
    dv = "0" if r == 11 else ("K" if r == 10 else str(r))
    return f"{n:,}".replace(",", ".") + "-" + dv

def fecha_aleatoria(inicio=date(1968,1,1), fin=date(2000,12,31)):
    delta = (fin - inicio).days
    return (inicio + timedelta(days=random.randint(0, delta))).strftime("%d/%m/%Y")

# ── Generar personas ─────────────────────────────────────────────────────────

personas = []
ruts_usados = set()
rut_base = 8_000_000

for i in range(47):
    sexo = random.choice(["M","F"])
    primer = random.choice(PRIMER_NOMBRE_M if sexo=="M" else PRIMER_NOMBRE_F)
    segundo = random.choice(SEGUNDO_NOMBRE)
    ap = random.choice(APELLIDOS)
    am = random.choice([a for a in APELLIDOS if a != ap])

    while rut_base in ruts_usados:
        rut_base += random.randint(120, 350)
    ruts_usados.add(rut_base)
    rut = rut_ficticio(rut_base)

    region = random.choice(list(REGIONES_COMUNAS.keys()))
    comuna = random.choice(REGIONES_COMUNAS[region])
    cargo  = random.choice(CARGOS)
    prof   = random.choice(PROFESIONES)
    exp_g  = random.randint(0, 25)
    exp_e  = random.randint(0, exp_g)
    calle  = random.choice(CALLES)
    num    = random.randint(100, 9999)
    tel    = f"+56 9 {random.randint(6000,9999)} {random.randint(1000,9999)}"
    correo = f"{primer.lower().replace('á','a').replace('é','e').replace('í','i').replace('ó','o').replace('ú','u')}.{ap.lower().replace('á','a').replace('é','e').replace('í','i').replace('ó','o').replace('ú','u')}@gmail.com"

    personas.append({
        "Apellido Paterno":           ap,
        "Apellido Materno":           am,
        "Primer Nombre":              primer,
        "Segundo Nombre":             segundo,
        "RUT":                        rut,
        "Fecha de Nacimiento":        fecha_aleatoria(),
        "Estado Civil":               random.choice(ESTADOS_CIVILES),
        "Nacionalidad":               random.choice(NACIONALIDADES),
        "Teléfono":                   tel,
        "Correo Electrónico":         correo,
        "Dirección":                  f"{calle} #{num}",
        "Comuna":                     comuna,
        "Región":                     region,
        "Cargo":                      cargo,
        "Profesión":                  prof,
        "Experiencia General (años)": exp_g,
        "Experiencia Específica (años)": exp_e,
        "Talla Overol":               random.choice(TALLAS_OVEROL),
        "Talla Zapatos":              random.choice(TALLAS_ZAPATOS),
        "Licencia Conducir":          random.choice(LICENCIAS),
        "Categoría":                  random.choice(CATEGORIAS),
        "Especialidad":               random.choice(ESPECIALIDADES),
        "Notas":                      "",
    })

# ── Escribir Excel ───────────────────────────────────────────────────────────

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Candidatos"

HEADERS = list(personas[0].keys())

# Estilo encabezado
header_fill = PatternFill("solid", fgColor="E30613")
header_font = Font(bold=True, color="FFFFFF", size=10)

for col, h in enumerate(HEADERS, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.fill   = header_fill
    cell.font   = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Datos
for row, p in enumerate(personas, 2):
    for col, h in enumerate(HEADERS, 1):
        ws.cell(row=row, column=col, value=p[h])

# Ancho de columnas
col_widths = [18,18,16,16,14,16,14,12,18,28,22,14,24,22,22,10,10,12,12,14,12,14,20]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

ws.row_dimensions[1].height = 35
ws.freeze_panes = "A2"

out = "candidatos_ejemplo.xlsx"
wb.save(out)
print(f"[OK] Generado: {out}  ({len(personas)} personas)")
